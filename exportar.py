import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date
import database as db
from config import FUENTE, COLOR_BG, COLOR_CARD, COLOR_TEXTO, COLOR_MUTED


def exportar_excel(ruta=None):
    """Genera un archivo Excel con resumen, movimientos y totales por cuenta."""

    if not ruta:
        hoy = date.today()
        ruta = f"movimientos_{hoy.year}_{hoy.month:02d}.xlsx"

    wb = openpyxl.Workbook()

    _hoja_resumen(wb)
    _hoja_movimientos(wb)
    _hoja_por_cuenta(wb)

    wb.save(ruta)
    return ruta


# ── Estilos ────────────────────────────────────────────────────────────────────

def _estilo_encabezado(celda, color="378ADD"):
    celda.font      = Font(bold=True, color="FFFFFF", size=11)
    celda.fill      = PatternFill("solid", fgColor=color)
    celda.alignment = Alignment(horizontal="center", vertical="center")

def _estilo_titulo(celda):
    celda.font      = Font(bold=True, size=13)
    celda.alignment = Alignment(horizontal="left")

def _color_fila(ws, fila, color):
    for col in range(1, ws.max_column + 1):
        ws.cell(row=fila, column=col).fill = PatternFill("solid", fgColor=color)


# ── Hoja 1: Resumen ────────────────────────────────────────────────────────────

def _hoja_resumen(wb):
    ws = wb.active
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20

    # Título
    ws["A1"] = "Resumen de cuentas"
    _estilo_titulo(ws["A1"])
    ws["A2"] = f"Generado: {date.today().isoformat()}"
    ws["A2"].font = Font(italic=True, color="888780")

    # Encabezados
    encabezados = ["Cuenta", "Tipo", "Sitio", "Saldo actual", "Propósito"]
    for i, enc in enumerate(encabezados, start=1):
        celda = ws.cell(row=4, column=i, value=enc)
        _estilo_encabezado(celda)
    ws.row_dimensions[4].height = 24

    # Datos
    cuentas = db.obtener_cuentas()
    tipos   = {"debito": "Débito", "credito": "Crédito",
               "ahorros": "Ahorros", "efectivo": "Efectivo", "otro": "Otro"}
    sitios  = {s["id"]: s["nombre"] for s in db.obtener_sitios()}

    total = 0
    for i, cuenta in enumerate(cuentas, start=5):
        saldo = db.calcular_saldo(cuenta["id"])
        total += saldo
        ws.cell(row=i, column=1, value=cuenta["nombre"])
        ws.cell(row=i, column=2, value=tipos.get(cuenta["tipo"], ""))
        ws.cell(row=i, column=3, value=sitios.get(cuenta["sitio_id"], "Sin sitio"))
        celda_saldo = ws.cell(row=i, column=4, value=saldo)
        celda_saldo.number_format = '"$"#,##0.00'
        celda_saldo.font = Font(
            color="1D9E75" if saldo >= 0 else "D85A30", bold=True)
        ws.cell(row=i, column=5, value=cuenta.get("proposito", ""))
        if i % 2 == 0:
            _color_fila(ws, i, "F5F4EF")

    # Total
    fila_total = len(cuentas) + 5
    ws.cell(row=fila_total, column=3, value="TOTAL").font = Font(bold=True)
    celda_total = ws.cell(row=fila_total, column=4, value=total)
    celda_total.number_format = '"$"#,##0.00'
    celda_total.font = Font(bold=True, size=12,
                            color="1D9E75" if total >= 0 else "D85A30")


# ── Hoja 2: Movimientos ────────────────────────────────────────────────────────

def _hoja_movimientos(wb):
    ws = wb.create_sheet("Movimientos")
    anchos = [12, 10, 14, 30, 18, 16, 14, 16, 16, 20]
    letras = ["A","B","C","D","E","F","G","H","I","J"]
    for l, a in zip(letras, anchos):
        ws.column_dimensions[l].width = a

    encabezados = ["Fecha", "Tipo", "Monto", "Descripción",
                   "Categoría", "Cuenta", "Tipo cuenta", "Sitio",
                   "Referencia", "Notas"]
    for i, enc in enumerate(encabezados, start=1):
        _estilo_encabezado(ws.cell(row=1, column=i, value=enc))
    ws.row_dimensions[1].height = 24

    movimientos = db.obtener_movimientos_completos()
    for i, mov in enumerate(movimientos, start=2):
        ws.cell(row=i, column=1, value=mov["fecha"])
        ws.cell(row=i, column=2, value="Ingreso" if mov["tipo"] == "ingreso" else "Gasto")
        celda_monto = ws.cell(row=i, column=3, value=mov["monto"])
        celda_monto.number_format = '"$"#,##0.00'
        celda_monto.font = Font(
            color="1D9E75" if mov["tipo"] == "ingreso" else "D85A30")
        ws.cell(row=i, column=4, value=mov["descripcion"])
        ws.cell(row=i, column=5, value=mov.get("categoria") or "")
        ws.cell(row=i, column=6, value=mov["cuenta"])
        ws.cell(row=i, column=7, value=mov["cuenta_tipo"])
        ws.cell(row=i, column=8, value=mov.get("sitio") or "Sin sitio")
        ws.cell(row=i, column=9, value=mov.get("referencia") or "")
        ws.cell(row=i, column=10, value=mov.get("notas") or "")
        if i % 2 == 0:
            _color_fila(ws, i, "F5F4EF")


# ── Hoja 3: Por cuenta ─────────────────────────────────────────────────────────

def _hoja_por_cuenta(wb):
    ws = wb.create_sheet("Por cuenta")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    encabezados = ["Cuenta", "Ingresos", "Gastos", "Saldo"]
    for i, enc in enumerate(encabezados, start=1):
        _estilo_encabezado(ws.cell(row=1, column=i, value=enc))
    ws.row_dimensions[1].height = 24

    hoy    = date.today()
    resumen = db.resumen_mes(hoy.year, hoy.month)

    for i, r in enumerate(resumen, start=2):
        saldo = r["total_ingresos"] - r["total_gastos"]
        ws.cell(row=i, column=1, value=r["nombre"])
        celda_ing = ws.cell(row=i, column=2, value=r["total_ingresos"])
        celda_ing.number_format = '"$"#,##0.00'
        celda_ing.font = Font(color="1D9E75")
        celda_gas = ws.cell(row=i, column=3, value=r["total_gastos"])
        celda_gas.number_format = '"$"#,##0.00'
        celda_gas.font = Font(color="D85A30")
        celda_sal = ws.cell(row=i, column=4, value=saldo)
        celda_sal.number_format = '"$"#,##0.00'
        celda_sal.font = Font(
            bold=True, color="1D9E75" if saldo >= 0 else "D85A30")
        if i % 2 == 0:
            _color_fila(ws, i, "F5F4EF")


if __name__ == "__main__":
    db.inicializar_db()
    ruta = exportar_excel()
    print(f"✅ Exportado: {ruta}")